import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List
from ratio.models.schemas import StatementMetadata, Transaction, ValidationSummary, RowValidationStatus

class ExcelGenerator:
    """
    Generates styled, intelligent Excel workbooks (.xlsx) with:
    - Summary Dashboard KPI Header Card.
    - Missing Page & Gap Alert banners (Red).
    - Color-coded transaction rows (Green = Valid, Yellow = Review, Red = Gap).
    - Tally & Audit ready formatting.
    """

    @classmethod
    def generate_workbook(
        cls,
        metadata: StatementMetadata,
        transactions: List[Transaction],
        validation: ValidationSummary
    ) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statement Audit"
        
        # Ensure grid lines are visible
        ws.views.sheetView[0].showGridLines = True
        
        # Color Palette (Financial Dark/Modern Theme)
        HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
        HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        VALID_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")  # Light Green
        REVIEW_FILL = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid") # Light Yellow
        GAP_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")    # Light Red
        
        GAP_BANNER_FILL = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid") # Red Alert
        GAP_BANNER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="0F172A")
        SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="64748B")
        KPI_LABEL_FONT = Font(name="Calibri", size=9, color="475569", bold=True)
        KPI_VALUE_FONT = Font(name="Calibri", size=12, color="0F172A", bold=True)
        
        THIN_BORDER = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # 1. Title Section
        ws["A1"] = f"RATIO — Financial Document Intelligence"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Bank: {metadata.institution} | File: {metadata.filename} | Total Pages: {metadata.total_pages}"
        ws["A2"].font = SUBTITLE_FONT

        # 2. KPI Summary Cards (Row 4 to 5)
        kpis = [
            ("TOTAL TRANSACTIONS", len(transactions), "A"),
            ("VALIDATED ROWS", validation.valid_rows, "B"),
            ("REVIEW NEEDED", validation.review_rows, "C"),
            ("GAPS DETECTED", len(validation.gaps), "D"),
            ("OPENING BALANCE", metadata.opening_balance or 0.0, "E"),
            ("CLOSING BALANCE", metadata.closing_balance or 0.0, "F"),
        ]
        
        for label, val, col in kpis:
            lbl_cell = ws[f"{col}4"]
            val_cell = ws[f"{col}5"]
            
            lbl_cell.value = label
            lbl_cell.font = KPI_LABEL_FONT
            
            if isinstance(val, float):
                val_cell.value = val
                val_cell.number_format = '₹#,##0.00'
            else:
                val_cell.value = val
            val_cell.font = KPI_VALUE_FONT
            
            lbl_cell.alignment = Alignment(horizontal="center")
            val_cell.alignment = Alignment(horizontal="center")
            
            lbl_cell.border = THIN_BORDER
            val_cell.border = THIN_BORDER

        start_row = 7

        # 3. Missing Page & Gap Alert Banner (if gaps detected)
        if validation.has_gaps:
            for gap in validation.gaps:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
                alert_cell = ws.cell(row=start_row, column=1)
                alert_cell.value = gap.message
                alert_cell.fill = GAP_BANNER_FILL
                alert_cell.font = GAP_BANNER_FONT
                alert_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[start_row].height = 28
                start_row += 1
            start_row += 1

        # 4. Table Headers
        headers = ["Txn Date", "Description", "Ref / Chq No", "Debit (₹)", "Credit (₹)", "Balance (₹)", "Validation Status & Notes"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        ws.row_dimensions[start_row].height = 24
        table_header_row = start_row
        start_row += 1

        # 5. Populate Transactions
        for trx in transactions:
            ws.cell(row=start_row, column=1, value=trx.date).alignment = Alignment(horizontal="center")
            ws.cell(row=start_row, column=2, value=trx.description)
            ws.cell(row=start_row, column=3, value=trx.reference or "-").alignment = Alignment(horizontal="center")
            
            # Debit
            c_deb = ws.cell(row=start_row, column=4, value=trx.debit)
            c_deb.number_format = '₹#,##0.00'
            c_deb.alignment = Alignment(horizontal="right")
            
            # Credit
            c_cred = ws.cell(row=start_row, column=5, value=trx.credit)
            c_cred.number_format = '₹#,##0.00'
            c_cred.alignment = Alignment(horizontal="right")
            
            # Balance
            c_bal = ws.cell(row=start_row, column=6, value=trx.balance)
            c_bal.number_format = '₹#,##0.00'
            c_bal.alignment = Alignment(horizontal="right")
            
            # Status & Notes
            status_cell = ws.cell(row=start_row, column=7, value=f"{trx.status.value}: {trx.validation_message or ''}")
            
            # Apply row formatting based on validation status
            row_fill = VALID_FILL
            if trx.status == RowValidationStatus.REVIEW_NEEDED:
                row_fill = REVIEW_FILL
            elif trx.status == RowValidationStatus.GAP_DETECTED or trx.status == RowValidationStatus.ERROR:
                row_fill = GAP_FILL
                
            for c_idx in range(1, 8):
                cell = ws.cell(row=start_row, column=c_idx)
                cell.fill = row_fill
                cell.border = THIN_BORDER

            start_row += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < table_header_row:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        ws.column_dimensions['B'].width = 36  # Description column wider
        ws.column_dimensions['G'].width = 50  # Validation Notes column wider

        # Auto filter on table
        ws.auto_filter.ref = f"A{table_header_row}:G{start_row - 1}"

        # Write to byte buffer
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @classmethod
    def generate_capital_gains_workbook(cls, metadata: StatementMetadata, cg_data) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Capital Gains (Schedule CG)"
        ws.views.sheetView[0].showGridLines = True

        # Sheet Header
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = "CAPITAL GAINS TAX COMPUTATION (ITR SCHEDULE CG READY)"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # KPI Summary Boxes
        kpis = [
            ("TOTAL PURCHASE COST (₹)", cg_data.total_purchase_cost, "A", "B"),
            ("TOTAL SALE PROCEEDS (₹)", cg_data.total_sale_value, "C", "D"),
            ("NET STCG (15% / 20%)", cg_data.total_stcg, "E", "F"),
            ("NET LTCG (12.5% / 10%)", cg_data.total_ltcg, "G", "H"),
        ]

        for label, val, c1, c2 in kpis:
            ws.merge_cells(f"{c1}3:{c2}3")
            ws.merge_cells(f"{c1}4:{c2}4")
            lbl_cell = ws[f"{c1}3"]
            val_cell = ws[f"{c1}4"]
            lbl_cell.value = label
            lbl_cell.font = Font(name="Calibri", size=9, bold=True, color="64748B")
            lbl_cell.alignment = Alignment(horizontal="center")
            val_cell.value = val
            val_cell.font = Font(name="Calibri", size=13, bold=True, color="0F172A")
            val_cell.number_format = "₹#,##0.00"
            val_cell.alignment = Alignment(horizontal="center")

        # Table Headers
        headers = ["Scheme Name / Investment", "Folio No", "Purchase Date", "Purchase Cost (₹)", "Sale Date", "Sale Proceeds (₹)", "Net STCG (₹)", "Net LTCG (₹)"]
        ws.row_dimensions[6].height = 24
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=c_idx, value=h)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if c_idx in [2,3,5] else "left", vertical="center")

        # Data Rows
        start_row = 7
        for item in cg_data.items:
            row_data = [
                item.scheme_name,
                item.folio_no or "-",
                item.purchase_date or "-",
                item.purchase_cost,
                item.sale_date or "-",
                item.sale_value,
                item.stcg,
                item.ltcg
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                if c_idx in [4, 6, 7, 8]:
                    cell.number_format = "₹#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif c_idx in [2, 3, 5]:
                    cell.alignment = Alignment(horizontal="center")
            start_row += 1

        # Auto column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        ws.column_dimensions['A'].width = 38

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
